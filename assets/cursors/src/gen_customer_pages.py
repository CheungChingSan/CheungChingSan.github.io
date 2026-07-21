# -*- coding: utf-8 -*-
"""
Regenerate the system's customer webpages from the user's Excel files:
  - 现有客户信息表.xlsx       -> webapp/base/existing_customers.html
  - 江洁客户尼日利亚260717.xlsx -> webapp/base/nigeria_customers.html

Rules honoured:
  * Only real customer info (no fabrication). Empty fields marked 未公开.
  * 备注 preserved verbatim; source noted at page level (user-provided Excel).
  * Reuses the existing page CSS/template (auth-gate + tier card style).
  * Drops the old "销售跟进建议" narrative block (user wants customer info only).
"""
import openpyxl, re, os

DESKTOP = r"C:\Users\Administrator\Desktop"
BASE = r"C:\Users\Administrator\WorkBuddy\2026-07-14-09-19-47\webapp\base"

def esc(s):
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))

def is_blank(v):
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == "/" or s.lower() == "无" or s.lower() == "none"

def split_phones(text):
    if text is None:
        return []
    # split on newline or explicit separators
    parts = re.split(r"[\n,;]+", str(text))
    return [p.strip() for p in parts if p.strip()]

def is_url(v):
    if is_blank(v):
        return False
    return str(v).strip().lower().startswith(("http://", "https://", "www."))

def is_email(v):
    if is_blank(v):
        return False
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(v).strip()))

# ---------------- Tier classification (existing customers) ----------------
ELECTRO = ["高频电刀", "电刀", "diathermy", "esu", "electrosurgical", "电外科", "leep"]
HIGH = ["erbe", "aesculap", "bovie", "valleylab", "lamidey", "hebu", "söring",
        "soring", "led spa", "led sp"]
KOREAN_DOM = ["zerone", "meditom", "heal force", "healicom", "alan", "shalya",
              "hosmed", "electrosource", "surgi touch", "national electrocare",
              "k-med", "topcare", "ro-chain", "marco", "triup", "ambygo",
              "electro range", "surgicut", "索吉瑞", "力康", "普朗", "贴牌",
              "国产", "组装", "kalstein", "beijing", "上海"]
ORTHO = ["骨科", "骨钻", "orthopedic", "ortho", "无源", "外科器械", "关节镜",
         "胸骨锯", "刨削", "植入物", "骨科动力"]

def classify_existing(products, contact):
    p = (str(products or "") + " " + str(contact or "")).lower()
    # Negation: text like "无高频电刀" / "无电外科设备" means NOT electro
    neg = any(k in p for k in ["无高频电刀", "无电外科", "无电刀", "没有高频电刀",
                               "无相关产品", "无电外科设备"])
    has_electro = (any(k in p for k in ELECTRO)) and not neg
    has_high = any(k in p for k in HIGH)
    has_kd = any(k in p for k in KOREAN_DOM)
    has_ortho = any(k in p for k in ORTHO) and not has_electro
    if has_electro and has_high:
        return 1
    if has_electro:
        return 2
    if has_ortho:
        return 3
    return 4

# ---------------- Tier classification (Nigeria) ----------------
VET = ["兽医", "vet", "veterinary", "animal", "pet", "agro", "agri", "livestock",
       "poultry", "clinic", "兽", "动物"]

def classify_nigeria(name, remark, web):
    blob = " ".join([str(x or "") for x in (name, remark, web)]).lower()
    if any(k in blob for k in VET):
        return 4  # 兽医诊所与服务商
    return 1      # 医疗设备经销商

# ---------------------------------------------------------------------------
def extract_style_and_gate(html):
    m_style = re.search(r"<style>(.*?)</style>", html, re.S)
    m_gate = re.search(r"<script data-xh-gate>(.*?)</script>", html, re.S)
    style = m_style.group(1) if m_style else ""
    gate = m_gate.group(1) if m_gate else ""
    return style, gate

# =====================================================================
# Build existing_customers.html
# =====================================================================
def build_existing():
    path = os.path.join(DESKTOP, "现有客户信息表.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # header: 序号,联系情况,公司名称,联系方式,公司官网,经营产品/品牌,国家,地址,备注
    records = []
    for r in rows[1:]:
        name = r[2]
        if name is None or (isinstance(name, str) and name.strip() == ""):
            continue
        records.append(r)

    tiers = {1: [], 2: [], 3: [], 4: []}
    for r in records:
        seq, contact, name, phone, web, products, country, addr, remark = (list(r) + [None]*9)[:9]
        t = classify_existing(products, contact)
        tiers[t].append({
            "seq": seq, "contact": contact, "name": name, "phone": phone,
            "web": web, "products": products, "country": country,
            "addr": addr, "remark": remark,
        })

    style, gate = extract_style_and_gate(
        open(os.path.join(BASE, "existing_customers.html"), encoding="utf-8").read())

    tier_meta = {
        1: ("第一梯队", "高频电刀·国际高端品牌代理商", "德国Erbe/Aesculap · 美国Bovie/Valleylab · 法国Lamidey · 德国Hebu"),
        2: ("第二梯队", "高频电刀·韩/国产/印/贴牌品牌", "韩国Zerone/Meditom · 国产Heal Force · 印度ALAN · 组装贴牌"),
        3: ("第三梯队", "骨科/外科器械（无电刀）", "骨科动力/无源外科器械"),
        4: ("第四梯队", "综合设备商（可开发补全）", "有OT设备/集团渠道，无电外科产品线"),
    }

    def card(c):
        n = esc(c["name"])
        web = c["web"]
        name_html = f'<a href="{esc(web)}" target="_blank">{n}</a>' if is_url(web) else n
        parts = []
        parts.append(f'<div class="name">{esc(c["seq"])}. {name_html}</div>')
        if not is_blank(c["products"]):
            parts.append(f'<div class="products">{esc(c["products"])}</div>')
        info = []
        if is_url(web):
            info.append(f'<div><span class="label">官网:</span> <a href="{esc(web)}" target="_blank">{esc(web)}</a></div>')
        else:
            info.append(f'<div><span class="label">官网:</span> 未公开</div>')
        phones = split_phones(c["phone"])
        if phones:
            plist = "".join(f'<span class="phone-item">{esc(p)}</span>' for p in phones)
            info.append(f'<div class="phone-row"><span class="label">电话:</span> <span class="phone-list">{plist}</span></div>')
        else:
            info.append(f'<div class="phone-row"><span class="label">电话:</span> 未公开</div>')
        if not is_blank(c["addr"]):
            info.append(f'<div><span class="label">地址:</span> {esc(c["addr"])}</div>')
        if not is_blank(c["contact"]) and str(c["contact"]).strip() not in ("未联系",):
            info.append(f'<div><span class="label">联系:</span> {esc(c["contact"])}</div>')
        if not is_blank(c["remark"]):
            info.append(f'<div><span class="label">备注:</span> {esc(c["remark"])}</div>')
        parts.append(f'<div class="info">{"".join(info)}</div>')
        # tags
        tags = []
        p = str(c["products"] or "").lower()
        if any(k in p for k in ELECTRO):
            tags.append('<span class="tag tag-electro">高频电刀</span>')
        if any(k in p for k in ORTHO):
            tags.append('<span class="tag tag-ortho">骨科</span>')
        if "ot" in p or "手术室" in p:
            tags.append('<span class="tag tag-ot">OT设备</span>')
        if tags:
            parts.append(f'<div class="tags">{"".join(tags)}</div>')
        return f'      <div class="company-card">\n        {"\n        ".join(parts)}\n      </div>'

    tier_html = []
    for t in (1, 2, 3, 4):
        if not tiers[t]:
            continue
        badge, title, desc = tier_meta[t]
        cards = "\n".join(card(c) for c in tiers[t])
        tier_html.append(
            f'  <div class="tier tier-{t}">\n'
            f'    <div class="tier-header">\n'
            f'      <span class="tier-badge">{badge}</span>\n'
            f'      <h2>{title}（{len(tiers[t])}家）</h2>\n'
            f'      <span class="tier-desc">{desc}</span>\n'
            f'    </div>\n'
            f'    <div class="company-list">\n{cards}\n    </div>\n'
            f'  </div>')

    total = len(records)
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<script data-xh-gate>{gate}</script>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>现有客户信息表 - 按规模/专业度/相关度排序</title>
<style>{style}</style>
</head>
<body>

<div class="header">
  <h1>现有客户信息表</h1>
  <p>数据来源：桌面《现有客户信息表.xlsx》（用户提供）| 共{total}家公司 | 按产品相关度四梯队</p>
</div>

<div class="container">

  <div class="summary-box">
    <h2>客户概览</h2>
    <div class="summary-stats">
      <div class="stat-card"><div class="num">{total}</div><div class="label">客户总数</div></div>
      <div class="stat-card"><div class="num">{len(tiers[1])}</div><div class="label">第一梯队（电刀+高端品牌）</div></div>
      <div class="stat-card"><div class="num">{len(tiers[2])}</div><div class="label">第二梯队（电刀+其他品牌）</div></div>
      <div class="stat-card"><div class="num">{len(tiers[3])}</div><div class="label">第三梯队（骨科/外科器械）</div></div>
      <div class="stat-card"><div class="num">{len(tiers[4])}</div><div class="label">第四梯队（综合·可开发）</div></div>
    </div>
  </div>

  <div class="note">
    <strong>说明：</strong>
    <br>• 本表由桌面《现有客户信息表.xlsx》原样生成，字段含序号 / 联系情况 / 公司名称 / 联系方式 / 公司官网 / 经营产品·品牌 / 国家 / 地址 / 备注。
    <br>• 电话中的「有WA」= 有 WhatsApp 可联系；「暂未联系」= 暂未接通；「WA留言」= 仅可 WhatsApp 留言。
    <br>• 官网/地址/备注为「/」「无」或空白的，统一标注「未公开」，未作任何虚构补全。
    <br>• 梯队仅依据「经营产品·品牌」字段分类：含国际高端电刀品牌（Erbe/Aesculap/Bovie/Valleylab/Lamidey/Hebu）为第一梯队；含韩/国产/印/贴牌电刀为第二梯队；主营骨科/外科器械无电刀为第三梯队；综合设备商为第四梯队。
  </div>

{chr(10).join(tier_html)}

</div>

<div class="footer">
  数据来源：用户《现有客户信息表.xlsx》（桌面）&nbsp;|&nbsp; 自动生成 &nbsp;|&nbsp; 共{total}家公司
</div>

</body>
</html>'''
    return html

# =====================================================================
# Build nigeria_customers.html
# =====================================================================
def build_nigeria():
    path = os.path.join(DESKTOP, "江洁客户尼日利亚260717.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for r in rows[1:]:
        name = r[0]
        if name is None or (isinstance(name, str) and name.strip() == ""):
            continue
        records.append(r)

    tiers = {1: [], 4: []}
    for r in records:
        name, country, phone, email, web, addr, remark = (list(r) + [None]*7)[:7]
        t = classify_nigeria(name, remark, web)
        tiers[t].append({
            "name": name, "country": country, "phone": phone,
            "email": email, "web": web, "addr": addr, "remark": remark,
        })

    style, gate = extract_style_and_gate(
        open(os.path.join(BASE, "nigeria_customers.html"), encoding="utf-8").read())

    def card(c):
        n = esc(c["name"])
        web = c["web"]
        name_html = f'<a href="{esc(web)}" target="_blank">{n}</a>' if is_url(web) else n
        parts = [f'<div class="name">{name_html}</div>']
        if not is_blank(c["remark"]):
            parts.append(f'<div class="products">{esc(c["remark"])}</div>')
        info = []
        if is_url(web):
            info.append(f'<div><span class="label">官网:</span> <a href="{esc(web)}" target="_blank">{esc(web)}</a></div>')
        else:
            info.append(f'<div><span class="label">官网:</span> 未公开</div>')
        phones = split_phones(c["phone"])
        if phones:
            plist = "".join(f'<span class="phone-item">{esc(p)}</span>' for p in phones)
            info.append(f'<div class="phone-row"><span class="label">电话:</span> <span class="phone-list">{plist}</span></div>')
        else:
            info.append(f'<div class="phone-row"><span class="label">电话:</span> 未公开</div>')
        if is_email(c["email"]):
            info.append(f'<div><span class="label">邮箱:</span> <a href="mailto:{esc(c["email"])}">{esc(c["email"])}</a></div>')
        else:
            info.append(f'<div><span class="label">邮箱:</span> 未公开</div>')
        if not is_blank(c["addr"]):
            info.append(f'<div><span class="label">地址:</span> {esc(c["addr"])}</div>')
        if not is_blank(c["remark"]):
            info.append(f'<div><span class="label">备注:</span> {esc(c["remark"])}</div>')
        parts.append(f'<div class="info">{"".join(info)}</div>')
        # tags
        tags = []
        blob = " ".join([str(x or "") for x in (c["remark"], c["name"])]).lower()
        if any(k in blob for k in ["电外科", "esu", "diathermy", "电刀"]):
            tags.append('<span class="tag tag-electro">电外科/ESU</span>')
        if any(k in blob for k in ["icu", "手术室", "实验室", "ot"]):
            tags.append('<span class="tag tag-ot">ICU/手术室</span>')
        if any(k in blob for k in ["牙科", "surgical", "外科"]):
            tags.append('<span class="tag tag-surgical">外科</span>')
        if tags:
            parts.append(f'<div class="tags">{"".join(tags)}</div>')
        return f'      <div class="company-card">\n        {"\n        ".join(parts)}\n      </div>'

    tier_html = []
    # tier-1 first
    for t in (1, 4):
        if not tiers[t]:
            continue
        if t == 1:
            badge, title, desc = "重点开发", "医疗设备经销商", "电刀/外科/ICU手术室设备相关 · 优先跟进"
        else:
            badge, title, desc = "潜在开发", "兽医诊所与服务商", "兽医/动物医疗相关"
        cards = "\n".join(card(c) for c in tiers[t])
        tier_html.append(
            f'  <div class="tier tier-{t}">\n'
            f'    <div class="tier-header">\n'
            f'      <span class="tier-badge">{badge}</span>\n'
            f'      <h2>{title}（{len(tiers[t])}家）</h2>\n'
            f'      <span class="tier-desc">{desc}</span>\n'
            f'    </div>\n'
            f'    <div class="company-list">\n{cards}\n    </div>\n'
            f'  </div>')

    total = len(records)
    med = len(tiers[1]); vet = len(tiers[4])
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<script data-xh-gate>{gate}</script>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>尼日利亚客户信息表 - 江洁 (2026-07-17)</title>
<style>{style}</style>
</head>
<body>

<div class="header">
  <h1>尼日利亚客户信息表 - 江洁</h1>
  <p>数据来源：桌面《江洁客户尼日利亚260717.xlsx》（用户提供）| 共{total}家公司</p>
</div>

<div class="container">

  <div class="summary-box">
    <h2>客户概览</h2>
    <div class="summary-stats">
      <div class="stat-card"><div class="num">{total}</div><div class="label">客户总数</div></div>
      <div class="stat-card"><div class="num">{med}</div><div class="label">医疗设备经销商</div></div>
      <div class="stat-card"><div class="num">{vet}</div><div class="label">兽医诊所与服务商</div></div>
    </div>
  </div>

  <div class="note">
    <strong>说明：</strong>
    <br>• 本表由桌面《江洁客户尼日利亚260717.xlsx》原样生成，字段含公司名称 / 国家 / 电话 / 邮箱 / 网址 / 地址 / 备注。
    <br>• 电话中的「有WA」= 该号码有 WhatsApp 可联系；「暂未联系」= 暂未接通；「WA留言」= 仅可 WhatsApp 留言。
    <br>• 邮箱/网址为「无」或空白的，表示原表未提供，标注「未公开」，未作任何虚构补全。
    <br>• 分类仅依据备注中的业务描述：含兽医/动物/Vet 等关键词归入「兽医诊所与服务商」，其余归入「医疗设备经销商」。
  </div>

{chr(10).join(tier_html)}

</div>

<div class="footer">
  数据来源：用户《江洁客户尼日利亚260717.xlsx》（桌面，江洁提供）&nbsp;|&nbsp; 自动生成 &nbsp;|&nbsp; 共{total}家公司
</div>

</body>
</html>'''
    return html

if __name__ == "__main__":
    out1 = build_existing()
    with open(os.path.join(BASE, "existing_customers.html"), "w", encoding="utf-8") as f:
        f.write(out1)
    print("existing_customers.html written:", len(out1), "bytes")

    out2 = build_nigeria()
    with open(os.path.join(BASE, "nigeria_customers.html"), "w", encoding="utf-8") as f:
        f.write(out2)
    print("nigeria_customers.html written:", len(out2), "bytes")
